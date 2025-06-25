import os
import re
import sys
import pandas as pd
import subprocess
import argparse
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# === CONFIGURATION ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
GAME_DATA_FOLDER = r"J:\\Steam\\steamapps\\common\\SovietRepublic\\media_soviet"
VEHICLE_SUBFOLDERS = ["vehicles", "trains", "airplanes", "ships", "helicopters", os.path.join("dlc2", "vehicles"), os.path.join("dlc1", "vehicles"), os.path.join("dlc3", "vehicles")]

BTF_TOOL_PATH = os.path.join(SCRIPT_DIR, "tools", "BTFTool.exe")
BTF_FILE = os.path.join(GAME_DATA_FOLDER, "sovietEnglish.btf")
BTF_TXT_PATH = os.path.join(GAME_DATA_FOLDER, "EN.txt")

# === STEP 1: Run BTFTool to export the btf to a txt file ===
def run_btftool(btf_tool_path, btf_file, output_txt):
    if not os.path.exists(btf_tool_path):
        raise FileNotFoundError(f"BTFTool.exe not found at: {btf_tool_path}")
    if not os.path.exists(btf_file):
        raise FileNotFoundError(f"BTF file not found: {btf_file}")

    cmd = [btf_tool_path, btf_file, f"--export={output_txt}"]
    subprocess.run(cmd, check=True)
    print(f"✅ BTFTool ran successfully, output: {output_txt}")

# === STEP 2: Load BTF name mapping ===
def load_btf_mapping(filepath):
    name_map = {}
    pattern = re.compile(r'^String\s+(\d+):\s+"(.*)"$')
    with open(filepath, "r", encoding="utf-8-sig") as file:
        for line in file:
            match = pattern.match(line.strip())
            if match:
                name_map[int(match.group(1))] = match.group(2).encode('utf-8').decode('unicode_escape')
    return name_map

# === STEP 3: Parse each vehicle's script.ini ===
def parse_script_ini(filepath, btf_name_map):
    data = {
        'Name': 'N/A',
        'StartYear': 'N/A',
        'EndYear': 'N/A',
        'Capacity': 'N/A',
        'CostRub': 'N/A',
        'CostUsd': 'N/A',
        'Speed': 'N/A',
        'TransportType': 'N/A',
        'Country': 'N/A',
        'VehicleType': 'N/A',
    }

    excluded = False
    resource_type = None
    skill_type = None
    skill_capacity = None

    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        lines = f.readlines()

    for line in lines:
        if line.startswith("$PURCHASE_EXCLUDE"):
            excluded = True
            break
        elif line.startswith("$NAME_STR"):
            match = re.search(r'\$NAME_STR\s+"(.*?)"', line)
            if match:
                data['Name'] = match.group(1)
        elif line.startswith("$NAME"):
            match = re.search(r'\$NAME\s+(\d+)', line)
            if match:
                name_id = int(match.group(1))
                data['Name'] = btf_name_map.get(name_id, f"UnknownNameID_{name_id}")
        elif line.startswith("$AVAILABLE"):
            match = re.search(r'\$AVAILABLE\s+(\d+)\s+(\d+)', line)
            if match:
                data['StartYear'], data['EndYear'] = match.groups()
        elif line.startswith("$RESOURCE_CAPACITY"):
            match = re.search(r'\$RESOURCE_CAPACITY\s+(\d+)', line)
            if match:
                data['Capacity'] = match.group(1)
        elif line.startswith("$COST_RUB"):
            match = re.search(r'\$COST_RUB\s+(\d+)', line)
            if match:
                data['CostRub'] = match.group(1)
        elif line.startswith("$COST_USD"):
            match = re.search(r'\$COST_USD\s+(\d+)', line)
            if match:
                data['CostUsd'] = match.group(1)
        elif line.startswith("$MOVEMENT_SPEED"):
            match = re.search(r'\$MOVEMENT_SPEED\s+(\d+)', line)
            if match:
                data['Speed'] = match.group(1)
        elif line.startswith("$RESOURCE_TRANSPORT_TYPE"):
            match = re.search(r'\$RESOURCE_TRANSPORT_TYPE\s+(\S+)', line)
            if match:
                cleaned = match.group(1).replace("RESOURCE_TRANSPORT_", "").replace("_", " ").lower()
                if cleaned == "passanger":
                    cleaned = "passenger"
                resource_type = cleaned
        elif "$SKILL_" in line:
            match = re.search(r'\$(SKILL_[A-Z_]+)\s+(\d+)', line)
            if match:
                skill = match.group(1).replace("SKILL_", "").replace("_", " ").lower()
                skill_type = skill
                skill_capacity = match.group(2)
        elif line.startswith("$TYPE"):
            match = re.search(r'\$TYPE\s+(VEHICLETYPE_[A-Z_]+)', line)
            if match:
                vtype_raw = match.group(1)
                vtype_clean = vtype_raw.replace("VEHICLETYPE_", "").replace("_", " ").lower()
                data['VehicleType'] = vtype_clean
                if vtype_clean == "helicopter":
                    sheet = "air"
                elif vtype_clean.startswith("rail"):
                    sheet = "rail"
                elif vtype_clean.startswith("road"):
                    sheet = "road"
                elif vtype_clean.startswith("ship"):
                    sheet = "water"
                elif vtype_clean.startswith("air"):
                    sheet = "air"
                else:
                    sheet = "other"
                data['_Sheet'] = sheet
        elif line.startswith("$COUNTRY"):
            match = re.search(r'\$COUNTRY\s+(\d+)', line)
            if match:
                country_id = int(match.group(1))
                data['Country'] = btf_name_map.get(country_id, f"UnknownCountryID_{country_id}")

    if excluded:
        return None

    if skill_type:
        data['TransportType'] = skill_type
        data['Capacity'] = skill_capacity
    elif resource_type:
        data['TransportType'] = resource_type

    return data

# === STEP 4: Gather vehicle data ===
def gather_all_vehicle_data(filter_year=None):
    run_btftool(BTF_TOOL_PATH, BTF_FILE, BTF_TXT_PATH)
    btf_name_map = load_btf_mapping(BTF_TXT_PATH)
    vehicles_by_sheet = defaultdict(list)

    for subfolder in VEHICLE_SUBFOLDERS:
        folder_path = os.path.join(GAME_DATA_FOLDER, subfolder)
        if not os.path.exists(folder_path):
            continue

        for vehicle_name in os.listdir(folder_path):
            vehicle_folder = os.path.join(folder_path, vehicle_name)
            script_ini_path = os.path.join(vehicle_folder, "script.ini")
            if not os.path.isfile(script_ini_path):
                continue

            vehicle_data = parse_script_ini(script_ini_path, btf_name_map)
            if not vehicle_data:
                continue

            if filter_year:
                try:
                    if not (int(vehicle_data['StartYear']) <= filter_year <= int(vehicle_data['EndYear'])):
                        continue
                except:
                    continue

            sheet_name = vehicle_data.get('_Sheet', 'other')
            vehicles_by_sheet[sheet_name].append(vehicle_data)

    return vehicles_by_sheet

# === STEP 5: Write to Excel ===
def autosize_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

def export_to_excel(vehicles_by_sheet, year=None):
    filename = f"vehicles_{year}.xlsx" if year else "vehicles.xlsx"
    filepath = os.path.join(SCRIPT_DIR, filename)

    sheet_order = ['road', 'rail', 'water', 'air']

    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for sheet_name in sheet_order:
                if sheet_name in vehicles_by_sheet:
                    df = pd.DataFrame(vehicles_by_sheet[sheet_name])
                    df = df.drop(columns=['_Sheet'], errors='ignore')  # Remove _Sheet column
                    df.sort_values(by=['VehicleType', 'TransportType', 'StartYear'], inplace=True, ignore_index=True)
                    df.to_excel(writer, index=False, sheet_name=sheet_name)

            for sheet_name in sorted(set(vehicles_by_sheet) - set(sheet_order)):
                df = pd.DataFrame(vehicles_by_sheet[sheet_name])
                df = df.drop(columns=['_Sheet'], errors='ignore')  # Remove _Sheet column
                df.sort_values(by=['VehicleType', 'TransportType', 'StartYear'], inplace=True, ignore_index=True)
                df.to_excel(writer, index=False, sheet_name=sheet_name)

            writer.book.template = False
        wb = load_workbook(filepath)
        for sheet in wb.worksheets:
            sheet.auto_filter.ref = sheet.dimensions
            autosize_columns(sheet)
        wb.save(filepath)
    except PermissionError:
        print(f"❌ Cannot write to '{filepath}'. Please close it if it's open.")
        return

    print(f"✅ Excel file written: {filepath}")


# === MAIN ===
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export vehicle data to Excel")
    parser.add_argument("--year", type=int, help="Filter vehicles available in a specific year (e.g., 1985)")
    args = parser.parse_args()

    all_data = gather_all_vehicle_data(filter_year=args.year)
    export_to_excel(all_data, year=args.year)